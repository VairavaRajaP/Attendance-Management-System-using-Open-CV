from openpyxl import Workbook
import json
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder


def export_to_xl():
    wb = Workbook()
    ws = wb.create_sheet("attendance", 0)
    ColumnDimension(ws, bestFit=True)
    ws.merge_cells(range_string='A1:A2')
    ws['A1'] = "Name"

    with open('attendance.json', 'r') as file:
        datas = json.load(file)

    i = 0
    row_number = 3
    col_number = 1
    each_dates = []
    coln = 2
    for data in datas.items():
        for items in data:
            if not i % 2 == 1:
                ws.cell(row_number, col_number, items.upper())
                row_number += 1
            else:
                for date in items.keys():
                    for each_date in ws['1']:
                        each_dates.append(each_date.value)

                    if date not in each_dates: # date_included:
                        # print(date)
                        # coln = len(ws['A'])+1
                        ws.cell(1, coln, date)
                        ws.merge_cells(start_row=1, start_column=coln, end_row=1, end_column=coln+1)
                        ws.cell(2, coln, 'In Time')
                        ws.cell(row_number-1, coln, str(items[date]['IN']))
                        ws.cell(2, coln+1, 'Out Time')
                        ws.cell(row_number-1, coln + 1, str(items[date]['OUT']))
                        coln += 2
                    else:
                        for each_date in ws['1']:
                            if each_date.value == date:
                                address_in = (each_date.coordinate[:-1]+str(row_number-1))
                                address_out_col = column_index_from_string(each_date.coordinate[:-1])
                                ws[address_in] = str(items[date]['IN'])
                                ws.cell(row_number-1, address_out_col+1, str(items[date]['OUT']))
            i += 1

    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=12)

    ws.column_dimensions = dim_holder

    rows = range(1, 3)
    columns = range(1, ws.max_column + 1)
    for row in rows:
        for col in columns:
            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row, col).font = Font(color='FF0000')

    try:
        wb.save('Attendance_format1.xlsx')
    except PermissionError:
        print('Excel file is already open. Please close the file and try again')

